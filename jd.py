from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import DesiredCapabilities

import time
import re
import argparse
import os
from urllib.parse import urlparse
import openpyxl
import requests

def get_options():
    # Set up Chrome WebDriver with options
    options = webdriver.ChromeOptions()
    # Disable the first run dialog and other similar popups
    # options.add_argument("--headless")
    options.add_experimental_option("debuggerAddress", "localhost:9014")
    # options.add_argument("--headless=new")
    options.add_argument("--no-first-run")
    options.add_argument("--no-default-browser-check")
    options.add_argument("--disable-infobars")
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--disable-gpu")
    options.page_load_strategy = 'eager'
    return options

# parser = argparse.ArgumentParser(description="JD.com Scraping Application")
# parser.add_argument('-b', '--brand', required=True, type=str, help="Input Brand Name")
# args = parser.parse_args()

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}

wb = openpyxl.Workbook()
ws = wb.worksheets[0]

def load_page(url):
    options = get_options()
    driver = webdriver.Chrome(options=options)

    item_list = []
    is_loaded = False
    while is_loaded == False:
        driver.get(url)
        time.sleep(10)

        for down_index in range(7):
            element = driver.find_element(By.XPATH, '/html/body')
            print("page down")
            element.send_keys(Keys.PAGE_DOWN)
            time.sleep(15)
        is_reload = False
        reload_element = None
        try:
            reload_element = driver.find_element(By.XPATH, '//*[@id="J_scroll_loading"]/span/a')
        except:
            reload_element = None
        while reload_element:
            print("reloading")
            is_reload = True
            driver.execute_script("arguments[0].click();", reload_element)
            time.sleep(5)
            try:
                reload_element = driver.find_element(By.XPATH, '//*[@id="J_scroll_loading"]/span/a')
            except:
                reload_element = None
        if is_reload:
            for down_index in range(5):
                element = driver.find_element(By.XPATH, '/html/body')
                print("page down again")
                element.send_keys(Keys.PAGE_DOWN)
                time.sleep(10)
        items = driver.find_elements(By.XPATH, '//*[@id="J_goodsList"]/ul/li')
        print(len(items))
        if len(items) != 60:
            print("not load all items")
        is_loaded = True
        for item_index in range(1, len(items) + 1):
            a_element = driver.find_element(By.XPATH, '//*[@id="J_goodsList"]/ul/li[{}]/div/div[1]/a'.format(item_index))
            href = a_element.get_attribute("href")
            item_id = re.findall('\d+', href)[0]
            price_element = driver.find_element(By.XPATH, '//*[@id="J_goodsList"]/ul/li[{}]/div/div[2]/strong/i'.format(item_index))
            price = price_element.text
            image_element = driver.find_element(By.XPATH, '//*[@id="J_goodsList"]/ul/li[{}]/div/div[1]/a/img'.format(item_index))
            image_src = None
            try:
                image_src = image_element.get_attribute("src")[0:-5]
            except:
                pass
            print(image_src)
            item_list.append([item_id, price, image_src])
    # print(item_list)
    item_list = [item for item in item_list if item[2]]
    return item_list
def main(brand):
    item_list = []
    options = get_options()
    g_driver = webdriver.Chrome(options=options)

    url = "https://search.jd.com/search?keyword=leather shoes&psort=5&wq=leather shoes&psort=5&ev=exbrand_{}%5E&cid3=6914".format(brand)
    g_driver.get(url)
    time.sleep(15)

    page_elements = g_driver.find_elements(By.XPATH, '//*[@id="J_bottomPage"]/span[1]/a')

    page_item_list = load_page(url)
    item_list.extend(page_item_list)

    print(len(page_elements))

    for page_number, page_element in enumerate(page_elements):
        print(page_number)
        if page_number > 6:
            break
        if page_number < 2:
            continue
        try:
            print(page_element.text.strip())
            page_int = int(page_element.text.strip())
            print(page_int)
        except:
            break
        g_driver.execute_script("arguments[0].click();", page_element)
        time.sleep(1)
        url = g_driver.current_url
        print(url)
        page_item_list = load_page(url)
        item_list.extend(page_item_list)


    # print(item_list)
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}

    for item_index, item in enumerate(item_list):
        cell_index = item_index * 5 + 1
        ws['A{}'.format(cell_index)] = item[0]
        ws['B{}'.format(cell_index)] = item[1]
        if item[2] is not None:
            response = requests.get(item[2], headers=headers)
            image_path = os.path.join(brand, 'img{}.jpg'.format(item_index))
            with open(image_path, 'wb') as out_file:
                out_file.write(response.content)
                time.sleep(0.5)
            try:
                img = openpyxl.drawing.image.Image(image_path)
                img.anchor = 'D{}'.format(cell_index)
                ws.add_image(img)
            except:
                pass
    wb.save('{}.xlsx'.format(brand))
if __name__ == "__main__":
    ecco = "爱步（ECCO）"
    belle = "百丽（Belle）"
    cartelo = "卡帝乐鳄鱼（CARTELO）"
    goldlion = "金利来（Goldlion）"
    bata = "拔佳（Bata）"
    satchi = "沙驰（SATCHI）"
    fuguinio = "富贵鸟（FUGUINIAO）"
    kisscat = "接吻猫（KISS%20CAT）"
    stNsat = "星期六（St%26Sat）"
    elle = "她（ELLE）"
    charles = "CHARLES%26KEITH"
    aokang = "奥康（Aokang）"
    # brands_list = [ecco, belle, cartelo, goldlion, bata, satchi, fuguinio, kisscat, stNsat, elle]
    brands_list = [aokang]

    for brand in brands_list:
        if not os.path.exists(brand):
            # Create the directory
            os.makedirs(brand)
            print("Directory created successfully!")
        main(brand)
        print(brand, " has been finished")
